# import module
import openpyxl

# load excel with its path
wrkbk = openpyxl.load_workbook("testfilter.xlsx")

sh = wrkbk.active
count = 0
# iterate through each cell
for i in range(5, sh.max_row + 1): #start at the first temperature data in excel
    #change data
    cell = sh.cell(row=i, column=2)
    value = int(cell.value)
    if (value > 7) and (count % 10 == 0):
        #print("Value before")
        #print(value)
        cell.value = value - 8
        #print("value after")
        #print(cell.value, end=" ")

    count = count+1
# save the file
wrkbk.save(filename="outputfilter.xlsx")
